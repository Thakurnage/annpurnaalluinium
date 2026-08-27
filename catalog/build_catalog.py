#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Premium SEO-optimized showcase catalog builder — v2 (Ultra-Premium Edition)"""
import base64, html, io, json, re, subprocess, sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps

ROOT = Path("/home/user")
ASSETS = ROOT / "catalog_assets"
OUT = ROOT / "catalog"
OUT.mkdir(exist_ok=True)
sys.path.insert(0, str(ROOT / "catalog"))
from products_data import P, CATS, BASE_FEATS  # noqa: E402

CAT_COLORS = {
    "UWC": "#38bdf8", "UWD": "#3b82f6", "ALW": "#22d3ee", "ALD": "#60a5fa",
    "ALP": "#2dd4bf", "GLZ": "#818cf8", "STF": "#8b5cf6", "MST": "#94a3b8",
    "SLS": "#9ca3af", "RLS": "#f59e0b", "MOD": "#f472b6",
}
CAT_ICONS = {
    "UWC": "⬚", "UWD": "▤", "ALW": "◫", "ALD": "▣", "ALP": "▥", "GLZ": "◈",
    "STF": "▦", "MST": "⛓", "SLS": "✦", "RLS": "⤢", "MOD": "▧",
}

SHOP = {
    "name": "Annapurna Aluminium & UPVC — Janakpur, Nepal",
    "brand": "Annapurna Alu & UPVC",
    "short": "Annapurna Aluminium & UPVC",
    "owner_name": "Prof. Nageshwar Thakur",
    "phone": "+977 9817658719",
    "phone2": "+977 9817667115",
    "phone_link": "tel:+9779817658719",
    "whatsapp": "9779817658719",
    "email": "",
    "pan": "616619779",
    "address": "Janakpur Dham-8, Murli Chowk (Airport Road), Dhanusha, Madhesh Province, Nepal",
    "address_short": "Janakpur Dham-8, Murli Chowk (Airport Road)",
    "domain": "https://your-shop-domain.com",
    "geo": {"lat": 26.7288, "lng": 85.9248},
    "city": "Janakpur Dham",
    "district": "Dhanusha",
    "province": "Madhesh Province",
    "areas": ["Janakpur Dham", "Dhanusha", "Mahottari (Jaleshwor)", "Sarlahi (Malangwa)",
              "Siraha (Lahan)", "Sindhuli (Bardibas)", "Udayapur (Gaighat)", "Rajbiraj",
              "Birgunj", "Hetauda", "Kathmandu", "Nepal-wide supply"],
}

OWNER_DIR = ASSETS / "owner"
OWNER_CACHE = {}
def owner_uri(name, width=400, q=75):
    """Embed owner image/avatar as data URI."""
    key = (name, width, q)
    if key in OWNER_CACHE:
        return OWNER_CACHE[key]
    p = OWNER_DIR / name
    if not p.exists():
        OWNER_CACHE[key] = None
        return None
    try:
        img = Image.open(p).convert("RGB")
        img = ImageOps.fit(img, (width, width), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, "PNG" if name.lower().endswith(".png") else "JPEG",
                 quality=q, optimize=True)
        uri = "data:image/%s;base64,%s" % ("png" if name.lower().endswith(".png") else "jpeg",
                                            base64.b64encode(buf.getvalue()).decode())
    except Exception:
        uri = None
    OWNER_CACHE[key] = uri
    return uri

# ---------------- render SVG art via node ----------------
def render_art():
    tpl = (ROOT / "catalog" / "templates.js").read_text(encoding="utf-8")
    m = re.search(r"var TEMPLATES = \(function \(\) \{(.*?)\}\)\(\);", tpl, re.S)
    body = m.group(1)
    js = (
        "var TEMPLATES = (function () {\n" + body + "\n})();\n"
        "function renderTemplate(key, params){ if(!TEMPLATES[key]) return ''; "
        "try { return TEMPLATES[key](params||{}); } catch(e){ return ''; } }\n"
        "var out = {};\n"
    )
    for pr in P:
        js += "out[%r] = renderTemplate(%r, %s);\n" % (pr["code"], pr["art"][0], json.dumps(pr["art"][1]))
    js += "process.stdout.write(JSON.stringify(out));"
    r = subprocess.run(["node", "-e", js], capture_output=True, text=True, timeout=60)
    if r.returncode != 0:
        print("NODE ERROR:", r.stderr[:2000]); sys.exit(1)
    return json.loads(r.stdout)

def svg_wrap(art):
    return ('<svg class="art" viewBox="0 0 200 160" preserveAspectRatio="xMidYMid meet" '
            'aria-hidden="true">' + art + "</svg>")

# ---------------- images ----------------
def data_uri(path, width=1200, q=80):
    img = Image.open(path).convert("RGB")
    if img.width > width:
        img = img.resize((width, int(img.height * width / img.width)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=q, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

PHOTO_CACHE = {}
def photo_uri(code, width=620, q=60):
    key = (code, width, q)
    if key in PHOTO_CACHE:
        return PHOTO_CACHE[key]
    p = ASSETS / "products" / ("%s.jpg" % code)
    if not p.exists():
        PHOTO_CACHE[key] = None
        return None
    try:
        img = Image.open(p).convert("RGB")
        img = ImageOps.fit(img, (width, int(width * 0.75)), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, "JPEG", quality=q, optimize=True)
        uri = "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
    except Exception:
        uri = None
    PHOTO_CACHE[key] = uri
    return uri

def esc(s):
    return html.escape(str(s), quote=True)

# ---------------- SEO helpers ----------------
def build_itemlist_jsonld():
    items = []
    for pr in P:
        items.append({
            "@type": "Product",
            "name": pr["name"],
            "description": pr["desc"],
            "sku": pr["code"],
            "brand": {"@type": "Brand", "name": SHOP["short"]},
            "url": SHOP["domain"] + "/#" + pr["code"].lower(),
            "image": SHOP["domain"] + "/images/products/" + pr["code"] + ".jpg",
        })
    return json.dumps({
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": "102 Premium Product Designs — Aluminium, Steel & UPVC Glass Works",
        "numberOfItems": len(items),
        "itemListElement": [{"@type": "ListItem", "position": i + 1, "item": it}
                            for i, it in enumerate(items)],
    }, ensure_ascii=False)

def build_faq_jsonld():
    faqs = [
        ("Do you provide service in Janakpur Dham and nearby areas (Dhanusha, Mahottari, Sarlahi)?",
         "Yes! We are located in Janakpur Dham and serve Dhanusha, Mahottari, Sarlahi, Siraha, Sindhuli, Udayapur — all of Madhesh Province with free site visit & measurement. Supply + installation is also available across the rest of Nepal."),
        ("What is the UPVC window price in Nepal?",
         "UPVC window price depends on size, profile quality (German/Chinese), double glazing and hardware. We provide premium quality at competitive rates in Nepal. Send us your window/door sizes for an exact quote — you will get the best price within 1 day."),
        ("UPVC window or aluminium window — which is better?",
         "UPVC windows offer better heat & sound insulation with multi-chamber profiles and zero maintenance — ideal for residential use. Aluminium windows are best for slim frames and high strength — perfect for commercial, shopfront and large openings. We provide custom solutions for both based on your requirement."),
        ("What is double glazing and what are its benefits?",
         "Double glazing means two glass panes with a sealed air/gas gap in between. It provides sound insulation, heat protection and condensation control — comfortable in both summer and winter. All our UPVC windows come with a double-glazed option — keeps your home cool even in Janakpur's heat."),
        ("Do you provide installation along with supply?",
         "Yes — we provide both supply and installation. Site visit, measurement, fabrication, delivery and professional installation — our team handles the entire process, along with after-sales service."),
    ]
    return json.dumps({
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [{"@type": "Question", "name": q,
                        "acceptedAnswer": {"@type": "Answer", "text": a}} for q, a in faqs],
    }, ensure_ascii=False)

def build_biz_jsonld():
    return json.dumps({
        "@context": "https://schema.org",
        "@type": "LocalBusiness",
        "@id": SHOP["domain"] + "/#business",
        "name": "Annapurna Aluminium & UPVC",
        "description": "Premium UPVC windows (jhyal), aluminium doors (dhoka), partitions, kitchen racks, steel railings & glass works in Janakpur Dham-8, Murli Chowk — Dhanusha, Madhesh Province, Nepal. 102+ unique designs, custom made, supply + installation.",
        "telephone": SHOP["phone"],
        "telephone2": SHOP["phone2"],
        "email": SHOP["email"] if SHOP["email"] else None,
        "taxID": SHOP["pan"],
        "vatID": SHOP["pan"],
        "founder": {"@type": "Person", "name": SHOP["owner_name"], "jobTitle": "Founder & Workshop Owner"},
        "image": SHOP["domain"] + "/images/og-cover.jpg",
        "logo": SHOP["domain"] + "/images/logo.png",
        "priceRange": "$$",
        "currenciesAccepted": "NPR",
        "paymentAccepted": "Cash, Bank Transfer, eSewa, Khalti, Mobile Banking",
        "openingHoursSpecification": [{
            "@type": "OpeningHoursSpecification",
            "dayOfWeek": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
            "opens": "09:00", "closes": "19:00",
        }],
        "address": {
            "@type": "PostalAddress",
            "streetAddress": "Janakpur Dham-8, Murli Chowk (Airport Road)",
            "addressLocality": "Janakpur Dham",
            "addressRegion": "Madhesh Province",
            "addressCountry": "NP",
        },
        "geo": {"@type": "GeoCoordinates", "latitude": SHOP["geo"]["lat"], "longitude": SHOP["geo"]["lng"]},
        "hasMap": "https://www.google.com/maps/search/?api=1&query=Murli+Chowk+Janakpur+Dham+Nepal",
        "areaServed": [
            {"@type": "City", "name": "Janakpur Dham"},
            {"@type": "AdministrativeArea", "name": "Dhanusha District"},
            {"@type": "AdministrativeArea", "name": "Mahottari District"},
            {"@type": "AdministrativeArea", "name": "Sarlahi District"},
            {"@type": "AdministrativeArea", "name": "Siraha District"},
            {"@type": "AdministrativeArea", "name": "Sindhuli District"},
            {"@type": "AdministrativeArea", "name": "Madhesh Province, Nepal"},
            {"@type": "Country", "name": "Nepal"},
        ],
        "knowsAbout": ["UPVC windows", "UPVC doors", "Aluminium windows", "Aluminium doors",
                       "Glass partitions", "Toughened glass", "Laminated glass", "Glass railings",
                       "Kitchen racks", "Steel railings", "Steel grills", "Shop front glazing",
                       "MS steel gates", "Stainless steel railings", "Spiral staircase",
                       "Modular kitchen", "Balcony enclosure", "Mosquito mesh", "Rolling shutter"],
        "url": SHOP["domain"],
        "sameAs": [],
    }, ensure_ascii=False)

def build_website_jsonld():
    return json.dumps({
        "@context": "https://schema.org",
        "@type": "WebSite",
        "name": SHOP["short"],
        "url": SHOP["domain"],
        "potentialAction": {"@type": "SearchAction",
                            "target": SHOP["domain"] + "/?q={search_term_string}",
                            "query-input": "required name=search_term_string"},
    }, ensure_ascii=False)

# ---------------- CSS (Ultra-Premium) ----------------
CSS = """
:root{
  --ink:#0b1220; --ink2:#111a2c; --ink3:#182338;
  --gold:#c9a227; --gold2:#f3c25b; --gold3:#8a6a1f;
  --paper:#f5f3ee; --card:#ffffff; --line:#e7e2d8; --mut:#5d6a7c;
  --serif:'Playfair Display',Georgia,'Times New Roman',serif;
  --sans:'Manrope','Segoe UI',system-ui,-apple-system,Roboto,Arial,sans-serif;
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:var(--sans);background:var(--paper);color:var(--ink);line-height:1.55;
  background-image:radial-gradient(circle at 15% 8%,rgba(201,162,39,.05),transparent 42%),
                   radial-gradient(circle at 85% 30%,rgba(17,26,44,.04),transparent 40%);}
::selection{background:var(--gold);color:#fff}
::-webkit-scrollbar{width:11px}
::-webkit-scrollbar-track{background:#e9e5db}
::-webkit-scrollbar-thumb{background:linear-gradient(var(--gold2),var(--gold3));border-radius:8px;border:2px solid #e9e5db}
a{color:inherit}
img{max-width:100%}
.skip{position:absolute;left:-999px;top:0;background:var(--gold);color:#0b1220;padding:10px 18px;
  font-weight:800;z-index:200;border-radius:0 0 10px 0;text-decoration:none}
.skip:focus{left:0}

/* ============ HERO ============ */
.hero{position:relative;min-height:100vh;display:flex;align-items:center;overflow:hidden;background:var(--ink)}
.hbg{position:absolute;inset:0;background-size:cover;background-position:center;
  opacity:0;transform:scale(1.04);transition:opacity 1.7s ease,transform 11s ease-out;will-change:opacity,transform}
.hbg.show{opacity:1;transform:scale(1.12)}
.hero::after{content:'';position:absolute;inset:0;
  background:linear-gradient(115deg,rgba(7,11,20,.93) 0%,rgba(7,11,20,.72) 42%,rgba(7,11,20,.35) 100%)}
.hprogress{position:absolute;left:0;right:0;bottom:0;height:3px;background:rgba(255,255,255,.14);z-index:6}
.hprogress i{display:block;height:100%;width:0;background:linear-gradient(90deg,var(--gold3),var(--gold2));box-shadow:0 0 12px var(--gold2)}
.heroinner{position:relative;z-index:4;width:min(1220px,92%);margin:0 auto;padding:110px 0 90px;color:#fff}
.kicker{display:inline-flex;align-items:center;gap:10px;letter-spacing:.3em;font-size:11px;font-weight:800;color:var(--gold2);
  border:1px solid rgba(243,194,91,.45);padding:8px 18px;border-radius:999px;margin-bottom:26px;text-transform:uppercase;
  background:rgba(201,162,39,.08);backdrop-filter:blur(4px)}
.kicker::before{content:'';width:26px;height:1px;background:var(--gold2)}
.hero h1{font-family:var(--serif);font-size:clamp(38px,6.2vw,76px);font-weight:700;line-height:1.06;letter-spacing:-.01em;text-wrap:balance}
.hero h1 .gold{background:linear-gradient(100deg,var(--gold2) 10%,#ffe9a8 45%,var(--gold) 90%);
  -webkit-background-clip:text;background-clip:text;color:transparent;font-style:italic}
.hero .sub{margin-top:22px;font-size:clamp(15px,1.9vw,18.5px);color:#c9d4e2;max-width:620px;line-height:1.7}
.hero .cta{display:flex;gap:14px;margin-top:36px;flex-wrap:wrap}
.btn-gold{background:linear-gradient(100deg,var(--gold),var(--gold2) 55%,var(--gold));
  background-size:200% 100%;color:#141005;font-weight:800;font-size:14px;letter-spacing:.04em;
  padding:15px 30px;border-radius:12px;text-decoration:none;box-shadow:0 10px 30px rgba(201,162,39,.35);
  transition:.35s;border:none;cursor:pointer}
.btn-gold:hover{background-position:100% 0;transform:translateY(-2px);box-shadow:0 16px 38px rgba(201,162,39,.45)}
.btn-ghost{background:rgba(255,255,255,.06);color:#fff;font-weight:700;font-size:14px;padding:15px 30px;border-radius:12px;
  text-decoration:none;border:1px solid rgba(255,255,255,.35);backdrop-filter:blur(6px);transition:.3s}
.btn-ghost:hover{background:rgba(255,255,255,.16);border-color:var(--gold2);color:var(--gold2)}
.herostats{display:flex;gap:clamp(22px,4vw,54px);margin-top:52px;flex-wrap:wrap;padding-top:30px;border-top:1px solid rgba(255,255,255,.14)}
.stat b{font-family:var(--serif);font-size:clamp(28px,3.6vw,44px);color:var(--gold2);display:block;line-height:1}
.stat span{font-size:11px;color:#93a3b8;letter-spacing:.18em;text-transform:uppercase;margin-top:6px;display:block}
.ownership{position:relative;z-index:4;width:min(1220px,92%);margin:-30px auto 0;display:flex;justify-content:flex-start;padding-bottom:10px}
.ownerchip{display:inline-flex;align-items:center;gap:14px;background:rgba(255,255,255,.08);border:1px solid rgba(243,194,91,.4);
  backdrop-filter:blur(10px);border-radius:16px;padding:10px 20px 10px 10px;box-shadow:0 14px 34px rgba(0,0,0,.35)}
.ownerchip img{width:56px;height:56px;border-radius:50%;border:2.5px solid var(--gold2);object-fit:cover;box-shadow:0 4px 14px rgba(201,162,39,.5)}
.ownerchip b{display:block;color:#fff;font-size:14.5px;font-family:var(--serif)}
.ownerchip span{display:block;color:var(--gold2);font-size:10.5px;letter-spacing:.14em;text-transform:uppercase;font-weight:700;margin-top:2px}
.ownerchip em{display:block;color:#b9c6d4;font-size:11.5px;font-style:italic;margin-top:3px}
.scrollhint{position:absolute;bottom:26px;left:50%;transform:translateX(-50%);z-index:5;color:rgba(255,255,255,.7);
  font-size:10px;letter-spacing:.3em;text-transform:uppercase;display:flex;flex-direction:column;align-items:center;gap:8px}
.scrollhint i{width:1px;height:38px;background:linear-gradient(var(--gold2),transparent);display:block;animation:drop 1.8s infinite}
@keyframes drop{0%{transform:scaleY(0);transform-origin:top}55%{transform:scaleY(1);transform-origin:top}100%{transform:scaleY(0);transform-origin:bottom}}

/* ============ NAV ============ */
.nav{position:sticky;top:0;z-index:90;background:rgba(11,18,32,.86);backdrop-filter:blur(14px);
  border-bottom:1px solid rgba(201,162,39,.25);box-shadow:0 6px 24px rgba(7,11,20,.25)}
.navwrap{width:min(1220px,94%);margin:0 auto;display:flex;align-items:center;gap:16px;padding:10px 0}
.brand{display:flex;align-items:center;gap:10px;color:#fff;text-decoration:none;flex-shrink:0}
.brand .logo{width:38px;height:38px;border-radius:10px;background:linear-gradient(135deg,var(--gold),var(--gold3));
  display:flex;align-items:center;justify-content:center;font-weight:800;color:#141005;font-size:14px;box-shadow:0 4px 14px rgba(201,162,39,.4);overflow:hidden}
.brand .logo img{width:100%;height:100%;object-fit:cover;border-radius:10px}
.brand b{font-family:var(--serif);font-size:16px;letter-spacing:.03em;color:#fff}
.brand small{display:block;font-size:8.5px;letter-spacing:.24em;color:var(--gold2);text-transform:uppercase;font-weight:700}
.navlinks{display:flex;gap:6px;overflow-x:auto;flex:1;scrollbar-width:none}
.navlinks::-webkit-scrollbar{display:none}
.navlinks a{color:#aeb9c9;text-decoration:none;font-size:11.5px;font-weight:700;white-space:nowrap;
  padding:8px 13px;border-radius:999px;transition:.25s;border:1px solid transparent}
.navlinks a:hover{color:#fff;background:rgba(255,255,255,.08)}
.navlinks a.active{color:#141005;background:linear-gradient(100deg,var(--gold),var(--gold2));border-color:transparent}
.nav-phone{flex-shrink:0;display:flex;align-items:center;gap:8px;color:var(--gold2);text-decoration:none;
  font-weight:800;font-size:13px;border:1px solid rgba(201,162,39,.5);padding:9px 16px;border-radius:999px;transition:.3s}
.nav-phone:hover{background:rgba(201,162,39,.15)}

/* ============ SECTIONS ============ */
.wrap{width:min(1220px,94%);margin:0 auto}
.catsec{padding:58px 0 10px;scroll-margin-top:70px;position:relative}
.cathead{display:flex;align-items:center;gap:18px;margin-bottom:26px;position:relative}
.catnum{font-family:var(--serif);font-style:italic;font-size:44px;line-height:1;color:transparent;
  -webkit-text-stroke:1px rgba(201,162,39,.55);flex-shrink:0}
.catico{width:52px;height:52px;border-radius:14px;background:var(--ink);color:var(--gold2);display:flex;
  align-items:center;justify-content:center;font-size:24px;box-shadow:0 8px 20px rgba(11,18,32,.25);flex-shrink:0}
.cathead h2{font-family:var(--serif);font-size:clamp(22px,3vw,30px);font-weight:700;letter-spacing:.01em}
.cathead h2 em{font-style:italic;color:var(--gold3)}
.cathead p{color:var(--mut);font-size:13px;margin-top:3px}
.cathead .count{margin-left:auto;background:linear-gradient(100deg,var(--gold),var(--gold2));color:#141005;
  font-weight:800;font-size:12px;padding:7px 15px;border-radius:999px;box-shadow:0 6px 16px rgba(201,162,39,.3);white-space:nowrap}

/* toolbar: search + filter */
.toolbar{width:min(1220px,94%);margin:26px auto 0;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.search{flex:1;min-width:240px;position:relative}
.search input{width:100%;padding:14px 18px 14px 46px;border-radius:14px;border:1.5px solid var(--line);
  background:var(--card);font-family:var(--sans);font-size:14px;color:var(--ink);outline:none;transition:.3s;box-shadow:0 4px 14px rgba(11,18,32,.05)}
.search input:focus{border-color:var(--gold);box-shadow:0 0 0 4px rgba(201,162,39,.15)}
.search svg{position:absolute;left:16px;top:50%;transform:translateY(-50%);color:#9aa7b8;pointer-events:none}
.filters{display:flex;gap:8px;flex-wrap:wrap}
.fchip{background:var(--card);border:1.5px solid var(--line);color:var(--mut);font-size:12px;font-weight:700;
  padding:9px 15px;border-radius:999px;cursor:pointer;transition:.25s;font-family:var(--sans)}
.fchip:hover{border-color:var(--gold);color:var(--gold3)}
.fchip.on{background:var(--ink);color:var(--gold2);border-color:var(--ink)}
.result-count{width:min(1220px,94%);margin:14px auto 0;font-size:12px;color:var(--mut);letter-spacing:.06em}

/* ============ CARDS ============ */
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(270px,1fr));gap:22px;margin-top:20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:18px;overflow:hidden;
  display:flex;flex-direction:column;box-shadow:0 5px 18px rgba(11,18,32,.06);transition:.35s;
  position:relative;animation:cardIn .5s both}
.card:hover{transform:translateY(-7px);box-shadow:0 24px 50px rgba(11,18,32,.18);border-color:rgba(201,162,39,.6)}
.card::after{content:'';position:absolute;inset:0 0 auto 0;height:3px;background:linear-gradient(90deg,var(--gold3),var(--gold2),var(--gold3));
  transform:scaleX(0);transform-origin:left;transition:transform .45s;z-index:2}
.card:hover::after{transform:scaleX(1)}
@keyframes cardIn{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:none}}
.artwrap{position:relative;background:linear-gradient(160deg,#ffffff,#eef0f5);border-bottom:1px solid var(--line);overflow:hidden;cursor:zoom-in}
.artwrap img.photo{display:block;width:100%;height:218px;object-fit:cover;transition:transform .8s ease;filter:saturate(1.02)}
.card:hover .artwrap img.photo{transform:scale(1.07)}
.artwrap .sheen{position:absolute;top:0;left:-80%;width:50%;height:100%;background:linear-gradient(105deg,transparent,rgba(255,255,255,.5),transparent);
  transform:skewX(-20deg);pointer-events:none}
.card:hover .sheen{animation:shine .9s ease}
@keyframes shine{to{left:130%}}
.artwrap svg.art{display:block;width:100%;height:auto;min-height:128px;padding:16px 14px 10px}
.specrow{display:flex;align-items:center;gap:10px;padding:7px 12px 9px;background:rgba(11,18,32,.03)}
.specrow svg.art{width:86px;min-height:0;padding:0}
.speclab{font-size:9px;color:var(--mut);letter-spacing:.16em;text-transform:uppercase;font-weight:800}
.badge{position:absolute;top:12px;left:12px;color:#fff;font-weight:800;font-size:10.5px;letter-spacing:.06em;
  padding:5px 10px;border-radius:8px;background:rgba(11,18,32,.78);backdrop-filter:blur(4px);box-shadow:0 4px 12px rgba(0,0,0,.25);z-index:3}
.mat{position:absolute;top:12px;right:12px;font-size:9.5px;font-weight:800;letter-spacing:.12em;
  background:rgba(255,255,255,.85);padding:4px 9px;border-radius:7px;backdrop-filter:blur(3px);z-index:3}
.cardbody{padding:17px 18px 18px;display:flex;flex-direction:column;gap:9px;flex:1}
.card h3{font-family:var(--serif);font-size:17.5px;font-weight:700;line-height:1.25}
.desc{color:var(--mut);font-size:12.5px;line-height:1.55}
.feats{list-style:none;display:flex;flex-direction:column;gap:5px;margin:2px 0 6px}
.feats li{font-size:12px;color:#3c4a5e;display:flex;gap:8px;align-items:flex-start;line-height:1.45}
.tick{flex:0 0 15px;height:15px;margin-top:1.5px;border-radius:50%;color:#141005;font-size:9.5px;font-weight:900;
  display:inline-flex;align-items:center;justify-content:center;background:linear-gradient(120deg,var(--gold2),var(--gold))}
.meta{margin-top:auto;display:flex;flex-direction:column;gap:8px;border-top:1px dashed var(--line);padding-top:12px}
.mblock b{font-size:9.5px;text-transform:uppercase;letter-spacing:.12em;color:var(--mut);display:block;margin-bottom:5px}
.chips{display:flex;flex-wrap:wrap;gap:5px}
.chip{background:#f0ede6;border:1px solid var(--line);color:#4a5568;font-size:10.5px;font-weight:600;padding:3px 9px;border-radius:999px}
.chip-fin{background:#fdf6e3;border-color:#ecd9a8;color:#8a6a1f}
.btn{display:block;text-align:center;text-decoration:none;font-weight:800;font-size:12.5px;letter-spacing:.03em;
  border:1.6px solid var(--ink);color:var(--ink);border-radius:11px;padding:10px;margin-top:4px;transition:.3s;position:relative;overflow:hidden}
.btn::before{content:'';position:absolute;inset:0;background:linear-gradient(100deg,var(--gold),var(--gold2));transform:translateY(101%);transition:.3s;z-index:-1}
.btn:hover{color:#141005;border-color:var(--gold)}
.btn:hover::before{transform:none}

/* reveal on scroll */
.reveal{opacity:0;transform:translateY(26px);transition:opacity .7s ease,transform .7s ease}
.reveal.in{opacity:1;transform:none}

/* ============ MEET THE OWNER ============ */
.meetsec{padding:56px 0 10px}
.meetcard{display:grid;grid-template-columns:1fr 1.05fr;gap:38px;align-items:center;background:var(--card);
  border:1px solid var(--line);border-radius:24px;overflow:hidden;box-shadow:0 18px 50px rgba(11,18,32,.1);
  position:relative}
.meetcard::before{content:'';position:absolute;inset:0 0 auto 0;height:4px;
  background:linear-gradient(90deg,var(--gold3),var(--gold2),var(--gold3))}
.meetimgs{position:relative;min-height:100%}
.meet-main{display:block;width:100%;height:100%;min-height:440px;object-fit:cover}
.meet-ch{position:absolute;bottom:18px;right:18px;width:118px;height:118px;border-radius:20px;
  object-fit:cover;border:3.5px solid var(--gold2);box-shadow:0 14px 34px rgba(0,0,0,.4);background:#fff}
.meetinfo{padding:34px 40px 38px 0}
.mkick{display:inline-flex;align-items:center;gap:8px;color:var(--gold3);font-size:11px;font-weight:800;
  letter-spacing:.24em;text-transform:uppercase;margin-bottom:12px}
.mkick::after{content:'';width:34px;height:1px;background:var(--gold)}
.meetinfo h2{font-family:var(--serif);font-size:clamp(28px,3.4vw,40px);font-weight:700;line-height:1.1}
.meetrole{color:var(--gold3);font-weight:700;font-size:13.5px;letter-spacing:.04em;margin-top:8px}
.meetbio{margin-top:16px;color:var(--mut);font-size:14.5px;line-height:1.8;font-style:italic}
.meetbio em{color:var(--gold3)}
.meetpts{margin-top:18px}
.meetcta{display:flex;gap:12px;margin-top:26px;flex-wrap:wrap}
.meetcta .btn-ghost{color:var(--ink);border-color:var(--ink);background:transparent}
.meetcta .btn-ghost:hover{background:var(--ink);color:var(--gold2)}
@media (max-width:860px){
  .meetcard{grid-template-columns:1fr;gap:0}
  .meet-main{min-height:300px}
  .meetinfo{padding:28px 22px 34px}
  .meet-ch{width:96px;height:96px}
}

/* ============ ABOUT SEO BLOCK ============ */
.aboutsec{background:linear-gradient(160deg,var(--ink) 0%,var(--ink2) 60%,#0d1526 100%);color:#dbe3ee;
  margin-top:70px;padding:70px 0;position:relative;overflow:hidden}
.aboutsec::before{content:'';position:absolute;inset:0;background:
  radial-gradient(600px 300px at 85% 0%,rgba(201,162,39,.14),transparent 60%)}
.aboutgrid{display:grid;grid-template-columns:1.1fr .9fr;gap:44px;align-items:center;position:relative;z-index:2}
.aboutsec h2{font-family:var(--serif);font-size:clamp(26px,3.4vw,40px);color:#fff;line-height:1.2;margin-bottom:18px}
.aboutsec h2 em{color:var(--gold2);font-style:italic}
.aboutsec p{color:#b7c3d4;font-size:14.5px;line-height:1.8;margin-bottom:14px}
.aboutpoints{list-style:none;display:flex;flex-direction:column;gap:12px;margin-top:22px}
.aboutpoints li{display:flex;gap:12px;align-items:flex-start;font-size:14px;font-weight:600;color:#e6ecf4}
.aboutpoints .ic{flex:0 0 34px;height:34px;border-radius:10px;background:rgba(201,162,39,.16);border:1px solid rgba(201,162,39,.4);
  display:flex;align-items:center;justify-content:center;color:var(--gold2);font-size:15px}
.abouter img{width:100%;border-radius:20px;box-shadow:0 30px 70px rgba(0,0,0,.45);border:1px solid rgba(201,162,39,.3)}
.ownercaption{position:absolute;left:16px;right:16px;bottom:16px;display:flex;align-items:center;gap:12px;
  background:rgba(11,18,32,.82);backdrop-filter:blur(10px);border:1px solid rgba(201,162,39,.45);
  border-radius:14px;padding:10px 14px}
.ownercaption img{width:46px;height:46px;border-radius:50%;border:2px solid var(--gold2);object-fit:cover}
.ownercaption b{display:block;color:#fff;font-family:var(--serif);font-size:14px}
.ownercaption span{display:block;color:var(--gold2);font-size:10px;letter-spacing:.12em;text-transform:uppercase;font-weight:700;margin-top:2px}
.abouter{position:relative}
.boardimg{width:100%;border-radius:16px;margin-top:14px;box-shadow:0 16px 40px rgba(0,0,0,.35);
  border:1px solid rgba(201,162,39,.4)}
.ownerbox{background:linear-gradient(160deg,rgba(201,162,39,.12),rgba(201,162,39,.04));border:1px solid rgba(201,162,39,.35);
  border-radius:16px;padding:18px}
.ownerrow{display:flex;gap:12px;align-items:center}
.ownerrow img{width:58px;height:58px;border-radius:50%;border:2.5px solid var(--gold2);object-fit:cover}
.ownerrow b{display:block;color:#fff;font-family:var(--serif);font-size:14.5px}
.ownerrow span{display:block;color:var(--gold2);font-size:10px;letter-spacing:.12em;text-transform:uppercase;font-weight:700;margin-top:2px}
.ownerrow a{color:#aeb9c9;font-size:12px;margin-top:4px}
.ownerrow a:hover{color:var(--gold2)}
.ownertag{font-style:italic;color:#c9d4e2;font-size:12.5px;margin-top:12px;padding-top:10px;border-top:1px dashed rgba(201,162,39,.35)}

/* ============ FAQ ============ */
.faqsec{padding:64px 0 20px}
.faqsec h2{font-family:var(--serif);text-align:center;font-size:clamp(26px,3.4vw,38px);margin-bottom:8px}
.faqsec .lead{text-align:center;color:var(--mut);font-size:14px;margin-bottom:30px}
.faqlist{max-width:860px;margin:0 auto;display:flex;flex-direction:column;gap:12px}
.faqitem{background:var(--card);border:1px solid var(--line);border-radius:14px;overflow:hidden;transition:.3s}
.faqitem:hover{border-color:rgba(201,162,39,.5)}
.faqitem summary{cursor:pointer;padding:17px 20px;font-weight:700;font-size:14.5px;list-style:none;display:flex;
  justify-content:space-between;align-items:center;gap:14px;color:var(--ink)}
.faqitem summary::-webkit-details-marker{display:none}
.faqitem summary .plus{flex-shrink:0;width:26px;height:26px;border-radius:50%;background:linear-gradient(120deg,var(--gold2),var(--gold));
  color:#141005;display:flex;align-items:center;justify-content:center;font-size:15px;font-weight:800;transition:.3s}
.faqitem[open] summary .plus{transform:rotate(45deg)}
.faqitem .ans{padding:0 20px 17px;color:var(--mut);font-size:13.5px;line-height:1.7}

/* ============ SERVICE AREAS ============ */
.areassec{padding:20px 0 10px}
.areachips{display:flex;flex-wrap:wrap;gap:10px;margin-top:6px}
.areachip{background:var(--card);border:1.5px solid var(--line);color:#3c4a5e;font-size:13px;font-weight:700;
  padding:10px 18px;border-radius:999px;transition:.3s;box-shadow:0 3px 10px rgba(11,18,32,.05)}
.areachip:hover{border-color:var(--gold);color:var(--gold3);transform:translateY(-2px);box-shadow:0 8px 18px rgba(201,162,39,.18)}
.areachip.main{background:linear-gradient(100deg,var(--gold),var(--gold2));color:#141005;border-color:transparent}
.areasnote{margin-top:18px;color:var(--mut);font-size:13.5px;line-height:1.8;background:var(--card);
  border:1px dashed rgba(201,162,39,.5);border-radius:14px;padding:16px 20px}

/* ============ BANNER ============ */
.banner{position:relative;margin:64px 0 0;min-height:340px;display:flex;align-items:center;overflow:hidden;background:var(--ink)}
.bbg{position:absolute;inset:0;background-size:cover;background-position:center;
  opacity:0;transform:scale(1.04);transition:opacity 1.6s ease,transform 10s ease-out}
.bbg.show{opacity:1;transform:scale(1.1)}
.banner::after{content:'';position:absolute;inset:0;background:linear-gradient(180deg,rgba(11,18,32,.45),rgba(11,18,32,.8))}
.banner .wrap{position:relative;z-index:2;color:#fff;padding:56px 0;text-align:center}
.banner h2{font-family:var(--serif);font-size:clamp(26px,4vw,44px);font-weight:700;line-height:1.25}
.banner h2 em{font-style:italic;color:var(--gold2)}
.banner p{color:#c9d4e2;max-width:600px;margin:14px auto 26px;line-height:1.7;font-size:14.5px}

/* ============ CONTACT ============ */
.contact{background:linear-gradient(170deg,#0b1220,#101a2e);color:#e8eef5;padding:64px 0 44px;position:relative;overflow:hidden}
.contact::before{content:'';position:absolute;inset:0;background:radial-gradient(700px 320px at 10% 100%,rgba(201,162,39,.12),transparent 60%)}
.contact .wrap{position:relative;z-index:2}
.contact .grid2{display:grid;grid-template-columns:repeat(auto-fit,minmax(230px,1fr));gap:28px}
.cbox h4{color:var(--gold2);font-size:11.5px;letter-spacing:.2em;text-transform:uppercase;margin-bottom:14px;display:flex;gap:9px;align-items:center}
.cbox h4::after{content:'';flex:1;height:1px;background:rgba(201,162,39,.3)}
.cbox p,.cbox a{color:#aeb9c9;font-size:13.5px;line-height:1.8;text-decoration:none;display:block}
.cbox a:hover{color:var(--gold2)}
.foot{margin-top:44px;text-align:center;color:#5f7084;font-size:12px;line-height:1.9;border-top:1px solid rgba(255,255,255,.08);padding-top:24px}
.foot b{color:#8fa0b4}

/* floaters */
.whatsapp{position:fixed;right:22px;bottom:22px;z-index:95;width:56px;height:56px;border-radius:50%;
  background:linear-gradient(135deg,#25d366,#128c7e);display:flex;align-items:center;justify-content:center;
  box-shadow:0 10px 30px rgba(18,140,126,.45);transition:.3s;text-decoration:none}
.whatsapp:hover{transform:scale(1.1) rotate(6deg)}
.whatsapp svg{width:28px;height:28px;fill:#fff}
.totop{position:fixed;right:26px;bottom:92px;z-index:95;width:44px;height:44px;border-radius:12px;
  background:var(--ink);color:var(--gold2);border:1px solid rgba(201,162,39,.5);display:flex;align-items:center;
  justify-content:center;cursor:pointer;opacity:0;pointer-events:none;transition:.35s;font-size:18px}
.totop.show{opacity:1;pointer-events:auto}
.totop:hover{background:var(--gold);color:#141005}

/* ============ LIGHTBOX ============ */
.lightbox{position:fixed;inset:0;z-index:200;background:rgba(7,11,20,.92);backdrop-filter:blur(8px);
  display:flex;flex-direction:column;align-items:center;justify-content:center;gap:16px;
  opacity:0;pointer-events:none;transition:.35s}
.lightbox.open{opacity:1;pointer-events:auto}
.lightbox img{max-width:min(900px,92vw);max-height:70vh;border-radius:14px;box-shadow:0 30px 80px rgba(0,0,0,.6);
  border:1px solid rgba(201,162,39,.4)}
.lightbox .lbmeta{color:#fff;text-align:center}
.lightbox .lbmeta b{font-family:var(--serif);font-size:22px;color:var(--gold2)}
.lightbox .lbmeta span{display:block;color:#93a3b8;font-size:12.5px;margin-top:4px;letter-spacing:.08em}
.lbclose{position:absolute;top:24px;right:28px;width:44px;height:44px;border-radius:50%;border:1px solid rgba(255,255,255,.3);
  background:rgba(255,255,255,.08);color:#fff;font-size:20px;cursor:pointer;transition:.3s;display:flex;align-items:center;justify-content:center}
.lbclose:hover{background:var(--gold);color:#141005;transform:rotate(90deg)}

@media (max-width:860px){
  .aboutgrid{grid-template-columns:1fr}
  .cathead .count{display:none}
  .hero{min-height:88vh}
  .nav-phone span{display:none}
  .nav-phone{padding:8px 12px}
}
@media print{
  .nav,.hero,.banner,.contact,.toolbar,.whatsapp,.totop,.faqsec,.aboutsec{display:none}
  .catsec{padding-top:16px}
  .card{break-inside:avoid}
}
@media (prefers-reduced-motion:reduce){
  *,*::before,*::after{animation-duration:.01ms!important;transition-duration:.01ms!important;scroll-behavior:auto!important}
}
"""

# ---------------- JS ----------------
JS = """
<script>
/* slideshow (hero + banner) */
(function(){
  var urls = @@HERO_URLS@@;
  var layers = document.querySelectorAll('#hero .hbg');
  var bar = document.getElementById('hbar');
  if(urls.length && layers.length>1){
    var DUR=3400, cur=0, i=0;
    layers[0].style.backgroundImage='url("'+urls[0]+'")';
    function tick(){
      bar.style.transition='none'; bar.style.width='0%';
      requestAnimationFrame(function(){ requestAnimationFrame(function(){
        bar.style.transition='width '+DUR+'ms linear'; bar.style.width='100%';
      });});
    }
    tick();
    setInterval(function(){
      i=(i+1)%urls.length;
      var inn=layers[1-cur], out=layers[cur];
      inn.style.backgroundImage='url("'+urls[i]+'")';
      inn.classList.add('show'); out.classList.remove('show');
      cur=1-cur; tick();
    },DUR);
  }
})();
(function(){
  var urls = @@BANNER_URLS@@;
  var layers = document.querySelectorAll('#banner .bbg');
  if(urls.length && layers.length>1){
    var cur=0, i=0;
    layers[0].style.backgroundImage='url("'+urls[0]+'")';
    setInterval(function(){
      i=(i+1)%urls.length;
      var inn=layers[1-cur], out=layers[cur];
      inn.style.backgroundImage='url("'+urls[i]+'")';
      inn.classList.add('show'); out.classList.remove('show');
      cur=1-cur;
    },4500);
  }
})();

/* scrollspy nav */
(function(){
  var links=[].slice.call(document.querySelectorAll('.navlinks a'));
  var secs=links.map(function(a){return document.querySelector(a.getAttribute('href'));}).filter(Boolean);
  window.addEventListener('scroll',function(){
    var pos=window.scrollY+120, cur=null;
    secs.forEach(function(s){ if(s.offsetTop<=pos) cur=s; });
    links.forEach(function(a){ a.classList.toggle('active', cur && a.getAttribute('href')==='#'+cur.id); });
    var tp=document.querySelector('.totop');
    if(tp) tp.classList.toggle('show', window.scrollY>700);
  },{passive:true});
})();

/* reveal on scroll */
(function(){
  var els=document.querySelectorAll('.reveal');
  if(!('IntersectionObserver' in window)){ els.forEach(function(e){e.classList.add('in');}); return; }
  var io=new IntersectionObserver(function(es){
    es.forEach(function(e){ if(e.isIntersecting){ e.target.classList.add('in'); io.unobserve(e.target); } });
  },{threshold:.08});
  els.forEach(function(e){ io.observe(e); });
})();

/* count-up stats */
(function(){
  var stats=document.querySelectorAll('.stat b[data-count]');
  if(!stats.length) return;
  function run(el){
    var target=parseInt(el.getAttribute('data-count'),10)||0, t0=null, dur=1600;
    function step(ts){
      if(!t0) t0=ts;
      var p=Math.min((ts-t0)/dur,1);
      var eased=1-Math.pow(1-p,3);
      el.textContent=Math.round(target*eased);
      if(p<1) requestAnimationFrame(step);
    }
    requestAnimationFrame(step);
  }
  if(!('IntersectionObserver' in window)){ stats.forEach(run); return; }
  var io=new IntersectionObserver(function(es){
    es.forEach(function(e){ if(e.isIntersecting){ run(e.target); io.unobserve(e.target); } });
  },{threshold:.4});
  stats.forEach(function(s){ io.observe(s); });
})();

/* live search + category filter */
(function(){
  var input=document.getElementById('q');
  var chips=[].slice.call(document.querySelectorAll('.fchip'));
  var cards=[].slice.call(document.querySelectorAll('.card'));
  var count=document.getElementById('rcount');
  function apply(){
    var q=(input?input.value:'').toLowerCase().trim();
    var cat=chips.filter(function(c){return c.classList.contains('on');}).map(function(c){return c.getAttribute('data-cat');});
    var n=0;
    cards.forEach(function(card){
      var hay=(card.getAttribute('data-search')||'').toLowerCase();
      var ok=(!q||hay.indexOf(q)>-1)&&(cat.length===0||cat.indexOf(card.getAttribute('data-cat'))>-1);
      card.style.display=ok?'':'none';
      if(ok) n++;
    });
    if(count) count.textContent=n+' of 102 designs shown';
  }
  if(input) input.addEventListener('input',apply);
  chips.forEach(function(c){
    c.addEventListener('click',function(){
      var on=c.classList.contains('on');
      if(c.getAttribute('data-cat')==='ALL'){ chips.forEach(function(x){x.classList.remove('on');}); }
      else{ c.classList.toggle('on'); chips.forEach(function(x){ if(x.getAttribute('data-cat')==='ALL') x.classList.remove('on'); }); }
      apply();
    });
  });
})();

/* lightbox */
(function(){
  var lb=document.getElementById('lb');
  var img=document.getElementById('lbimg');
  var meta=document.getElementById('lbmeta');
  document.querySelectorAll('.artwrap img.photo').forEach(function(im){
    im.addEventListener('click',function(){
      var card=im.closest('.card');
      img.src=im.src;
      meta.innerHTML='<b>'+card.querySelector('h3').textContent+'</b><span>'+card.getAttribute('data-cat')+' · '+card.querySelector('.badge').textContent+'</span>';
      lb.classList.add('open');
      document.body.style.overflow='hidden';
    });
  });
  function close(){ lb.classList.remove('open'); document.body.style.overflow=''; }
  document.getElementById('lbclose').addEventListener('click',close);
  lb.addEventListener('click',function(e){ if(e.target===lb) close(); });
  document.addEventListener('keydown',function(e){ if(e.key==='Escape') close(); });
})();

/* back to top */
document.querySelector('.totop').addEventListener('click',function(){ window.scrollTo({top:0,behavior:'smooth'}); });
</script>
"""

# ---------------- build HTML ----------------
def build_html(art_map):
    hero_urls = [u for p in P if (u := photo_uri(p["code"], 700, 50))]
    banner_urls = hero_urls[::9][:12] if len(hero_urls) > 12 else hero_urls

    prod_json = json.dumps([{"c": p["code"], "n": p["name"], "cat": p["cat"]} for p in P])

    cards = []
    for idx, pr in enumerate(P):
        color = CAT_COLORS[pr["cat"]]
        feats = "".join(
            '<li><span class="tick">&#10003;</span>%s</li>' % esc(f) for f in pr["extras"][:3])
        sizes = "".join('<span class="chip">%s</span>' % esc(s) for s in pr["sizes"])
        fins = "".join('<span class="chip chip-fin">%s</span>' % esc(f) for f in pr["finishes"])
        photo = photo_uri(pr["code"], 620, 58)
        alt = "%s — %s premium design, custom size" % (esc(pr["name"]), esc(pr["cat"]))
        if photo:
            visual = ('<figure class="artwrap" data-src="%s">'
                      '<img class="photo" src="%s" alt="%s" loading="lazy" decoding="async" width="620" height="465">'
                      '<span class="sheen"></span>'
                      '<div class="specrow">%s<span class="speclab">Technical View</span></div>'
                      '</figure>') % (photo, photo, alt, svg_wrap(art_map[pr["code"]]))
        else:
            visual = '<figure class="artwrap">' + svg_wrap(art_map[pr["code"]]) + "</figure>"
        cards.append(f"""
        <article class="card reveal" id="{esc(pr['code'].lower())}" data-cat="{esc(pr['cat'])}"
                 data-search="{esc(pr['code'] + ' ' + pr['name'] + ' ' + pr['cat'])}" itemscope itemtype="https://schema.org/Product">
          {visual}
          <span class="badge" itemprop="sku">{esc(pr['code'])}</span>
          <span class="mat" style="color:{color}">{esc(pr['cat'])}</span>
          <div class="cardbody">
            <h3 itemprop="name">{esc(pr['name'])}</h3>
            <p class="desc" itemprop="description">{esc(pr['desc'])}</p>
            <ul class="feats">{feats}</ul>
            <div class="meta">
              <div class="mblock"><b>Standard Sizes</b><div class="chips">{sizes}</div></div>
              <div class="mblock"><b>Finishes</b><div class="chips">{fins}</div></div>
            </div>
            <a class="btn" href="#contact">Enquire &#10148;</a>
          </div>
        </article>""")

    by_cat = {}
    for pr in P:
        by_cat.setdefault(pr["cat"], []).append(pr)

    sections = []
    for n, c in enumerate(CATS, 1):
        color = CAT_COLORS[c["key"]]
        items = by_cat[c["key"]]
        card_html = "".join(cards[P.index(p)] for p in items)
        sections.append(f"""
        <section class="catsec" id="sec-{esc(c['key'].lower())}" aria-labelledby="h-{esc(c['key'].lower())}">
          <div class="cathead reveal">
            <span class="catnum">{n:02d}</span>
            <span class="catico">{CAT_ICONS[c['key']]}</span>
            <div>
              <h2 id="h-{esc(c['key'].lower())}">{esc(c['name'])} <em>Collection</em></h2>
              <p>{esc(c['tag'])}</p>
            </div>
            <span class="count">{len(items)} premium designs</span>
          </div>
          <div class="grid">{card_html}</div>
        </section>""")

    nav = "".join(
        '<a href="#sec-%s">%s</a>' % (esc(c["key"].lower()), esc(c["name"])) for c in CATS)

    faq_html = "".join(
        '<details class="faqitem"><summary>%s<span class="plus">+</span></summary><div class="ans"><p>%s</p></div></details>'
        % (esc(q), esc(a)) for q, a in [
            ("Do you provide service in Janakpur Dham and nearby areas (Dhanusha, Mahottari, Sarlahi)?",
             "Yes! We are located in Janakpur Dham and serve Dhanusha, Mahottari, Sarlahi, Siraha, Sindhuli, Udayapur — all of Madhesh Province with free site visit & measurement. Supply + installation is also available across the rest of Nepal."),
            ("What is the UPVC window price in Nepal?",
             "UPVC window price depends on size, profile quality (German/Chinese), double glazing and hardware. We provide premium quality at competitive rates in Nepal. Send us your window/door sizes for an exact quote — you will get the best price within 1 day."),
            ("UPVC window or aluminium window — which is better?",
             "UPVC windows offer better heat & sound insulation with multi-chamber profiles and zero maintenance — ideal for residential use. Aluminium windows are best for slim frames and high strength — perfect for commercial, shopfront and large openings. We provide custom solutions for both based on your requirement."),
            ("Do you make custom sizes and custom designs?",
             "Absolutely. Every product is 100% custom-made, factory-finished to your exact measurements. Design consultation and free site measurement are available — you can also visit our showroom in Janakpur Dham to see 102+ designs."),
            ("What is double glazing and what are its benefits?",
             "Double glazing means two glass panes with a sealed air/gas gap in between. It provides sound insulation, heat protection and condensation control — comfortable in both summer and winter. All our UPVC windows come with a double-glazed option — keeps your home cool even in Janakpur's heat."),
            ("What warranty do you offer on products?",
             "We use premium profiles and branded hardware. Warranty-backed guarantee is provided on profiles, glass and hardware — contact us for details."),
            ("Do you provide installation along with supply?",
             "Yes — we provide both supply and installation. Site visit, measurement, fabrication, delivery and professional installation — our team handles the entire process, along with after-sales service."),
        ])

    fonts_css = (ROOT / "catalog" / "font_embed.css").read_text(encoding="utf-8") if (ROOT / "catalog" / "font_embed.css").exists() else ""

    # Avatar-based favicon (small circular owner photo)
    fav_av = owner_uri("avatar1_circle.png", 96, 80)
    favicon = fav_av if fav_av else ("data:image/svg+xml," + urllib_quote('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 64 64">'
        '<rect width="64" height="64" rx="14" fill="#0b1220"/>'
        '<rect x="10" y="14" width="44" height="36" rx="3" fill="none" stroke="#f3c25b" stroke-width="4"/>'
        '<line x1="32" y1="14" x2="32" y2="50" stroke="#f3c25b" stroke-width="3"/>'
        '<line x1="10" y1="50" x2="54" y2="50" stroke="#f3c25b" stroke-width="3"/></svg>'))

    doc = """<!DOCTYPE html>
<html lang="en" dir="ltr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>UPVC Windows &amp; Aluminium Doors in Janakpur, Nepal | Steel &amp; Glass Works — 102 Premium Designs</title>
<meta name="description" content="Best UPVC windows, aluminium doors, toughened glass & steel works in Janakpur Dham, Dhanusha — Madhesh Province, Nepal. 102+ premium custom designs: UPVC windows & doors, aluminium partitions, glass railings, MS gates, SS railing, shop front glazing, ACP cladding. Free measurement, factory-finished, supply + installation — Nepal-wide.">
<meta name="keywords" content="UPVC window Janakpur, UPVC window Nepal, UPVC door manufacturer Nepal, aluminium door Janakpur, aluminium window Nepal, glass works Janakpur Dham, toughened glass railing Nepal, glass partition Nepal, MS steel gate Janakpur, stainless steel railing Nepal, shop front glazing Nepal, ACP cladding Nepal, balcony glass railing Nepal, UPVC window price Nepal, aluminium door price Nepal, window grill Janakpur, glass railing Dhanusha, aluminium partition Madhesh, UPVC window Kathmandu, steel fabrication Janakpur, modular kitchen Nepal, rolling shutter Nepal, mosquito mesh window Nepal, spiral staircase Nepal, Janakpur Dham construction, Dhanusha glass shop">
<meta name="robots" content="index, follow, max-image-preview:large">
<meta name="author" content="@@SHOP_NAME@@">
<meta name="theme-color" content="#0b1220">
<meta name="geo.region" content="NP-P2">
<meta name="geo.placename" content="Janakpur Dham, Dhanusha, Nepal">
<meta name="geo.position" content="26.7288;85.9248">
<meta name="ICBM" content="26.7288, 85.9248">
<meta name="city" content="Janakpur Dham">
<meta name="district" content="Dhanusha">
<meta name="province" content="Madhesh Province">
<meta name="country" content="Nepal">
<link rel="canonical" href="@@DOMAIN@@/">
<link rel="alternate" hreflang="en" href="@@DOMAIN@@/">
<link rel="alternate" hreflang="ne" href="@@DOMAIN@@/ne/">
<link rel="alternate" hreflang="x-default" href="@@DOMAIN@@/">
<link rel="icon" href="@@FAVICON@@">
<meta property="og:type" content="website">
<meta property="og:site_name" content="@@SHOP_NAME@@">
<meta property="og:title" content="UPVC Windows &amp; Aluminium Doors in Janakpur, Nepal — 102 Premium Custom Designs">
<meta property="og:description" content="Windows, doors, partitions, glass solutions, storefronts, steel & stainless works, railings aur modular systems — Janakpur Dham, Dhanusha, Nepal. 102+ unique premium designs, 100% custom made, free measurement.">
<meta property="og:url" content="@@DOMAIN@@/">
<meta property="og:image" content="@@DOMAIN@@/images/og-cover.jpg">
<meta property="og:locale" content="en_US">
<meta property="og:locale:alternate" content="ne_NP">
<meta name="twitter:card" content="summary_large_image">
<meta name="twitter:title" content="UPVC Windows & Aluminium Doors in Janakpur, Nepal — 102 Premium Designs">
<meta name="twitter:description" content="102+ unique premium designs in Janakpur Dham — custom sizes, premium finishes, free measurement, supply + installation.">
<meta name="twitter:image" content="@@DOMAIN@@/images/og-cover.jpg">
<style>@@FONTS@@
@@CSS@@</style>
<script type="application/ld+json">@@BIZ_JSON@@</script>
<script type="application/ld+json">@@ITEMLIST_JSON@@</script>
<script type="application/ld+json">@@FAQ_JSON@@</script>
<script type="application/ld+json">@@WEBSITE_JSON@@</script>
</head>
<body>
<a class="skip" href="#main">Skip to catalog</a>

<header class="hero" id="hero">
  <div class="hbg show" id="hbg0" role="img" aria-label="Premium product showcase"></div>
  <div class="hbg" id="hbg1" role="img" aria-hidden="true"></div>
  <div class="hprogress"><i id="hbar"></i></div>
  <div class="heroinner">
    <span class="kicker reveal in">Premium Showcase Collection — 102 Designs</span>
    <h1 class="reveal in">Aluminium • Steel • UPVC<br><span class="gold">Glass Works</span> &amp; Designs</h1>
    <p class="sub reveal in">102+ unique, premium-quality product designs — UPVC windows &amp; doors, aluminium doors &amp; partitions, toughened glass, steel &amp; stainless works, railings and modular systems. <strong style="color:#fff">Your trusted workshop in Janakpur Dham, Dhanusha — Madhesh Province, Nepal:</strong> every product is custom-made to your exact measurements, with supply &amp; installation included.</p>
    <div class="cta reveal in">
      <a class="btn-gold" href="#main">Explore 102 Designs &#8595;</a>
      <a class="btn-ghost" href="#contact">Get Free Quote</a>
    </div>
    <div class="herostats reveal in">
      <div class="stat"><b data-count="102">0</b><span>Unique Designs</span></div>
      <div class="stat"><b data-count="11">0</b><span>Categories</span></div>
      <div class="stat"><b data-count="100">0</b><span>% Custom Made</span></div>
      <div class="stat"><b data-count="10">0</b><span>Year Warranty*</span></div>
    </div>
    <div class="ownerchip reveal in">
      <img src="@@AVATAR_GOLD@@" alt="Shop owner portrait — @@OWNER_NAME@@" width="72" height="72">
      <div><b>@@OWNER_NAME@@</b><span>Founder &amp; Workshop Owner — Janakpur Dham</span><em>"Quality is our promise — every product is crafted with heart."</em></div>
    </div>
  </div>
  <div class="scrollhint">Scroll<i></i></div>
</header>

<nav class="nav" aria-label="Categories">
  <div class="navwrap">
    <a class="brand" href="#hero" aria-label="Home">
      <span class="logo"><img src="@@NAV_AVATAR@@" alt="@@OWNER_NAME@@ — shop owner avatar" width="34" height="34"></span>
      <span><b>@@BRAND@@</b><small>Aluminium • Steel • UPVC Works</small></span>
    </a>
    <div class="navlinks">@@NAV@@</div>
    <a class="nav-phone" href="@@PHONE_LINK@@">&#9742; <span>@@PHONE@@</span></a>
  </div>
</nav>

<section class="meetsec" id="owner" aria-label="Meet the owner">
  <div class="wrap meetcard reveal">
    <div class="meetimgs">
      <img class="meet-main" src="@@MEET_IMG@@" alt="@@OWNER_NAME@@ — Founder and owner, premium workshop portrait" loading="lazy" decoding="async">
      <img class="meet-ch" src="@@CHAR_IMG@@" alt="@@OWNER_NAME@@ — brand character avatar" loading="lazy" decoding="async">
    </div>
    <div class="meetinfo">
      <span class="mkick">&#10022; Meet the Owner</span>
      <h2>@@OWNER_NAME@@</h2>
      <p class="meetrole">Founder &amp; Workshop Owner — Annapurna Aluminium &amp; UPVC, Janakpur Dham, Dhanusha, Nepal</p>
      <p class="meetbio">"I make every product as if it were for my own home — <em>no compromise on quality, measurement or finish</em>. Customers from Janakpur and all around Madhesh Province have trusted us for years — come and create the design of your dreams with us."</p>
      <ul class="aboutpoints meetpts">
        <li><span class="ic">&#10003;</span>Aluminium &amp; UPVC windows (jhyal), doors (dhoka), partitions &amp; kitchen racks</li>
        <li><span class="ic">&#10003;</span>Steel railings, grills &amp; fabrication — complete custom work</li>
        <li><span class="ic">&#10003;</span>German-profile UPVC, branded hardware, safety glass</li>
        <li><span class="ic">&#10003;</span>Personal quality check on every single delivery</li>
      </ul>
      <div class="meetcta">
        <a class="btn-gold" href="@@PHONE_LINK@@">&#9742; Call @@PHONE@@</a>
        <a class="btn-gold" href="@@PHONE_LINK2@@">&#9742; @@PHONE2@@</a>
        <a class="btn-ghost" href="https://wa.me/@@WHATSAPP@@" target="_blank" rel="noopener">WhatsApp Message</a>
      </div>
    </div>
  </div>
</section>

<div class="toolbar" role="search">
  <div class="search">
    <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.4"><circle cx="11" cy="11" r="7"/><line x1="21" y1="21" x2="16.5" y2="16.5"/></svg>
    <input type="search" id="q" placeholder="Search designs — e.g. 'sliding', 'railing', 'UWC-01'..." aria-label="Search products">
  </div>
  <div class="filters" aria-label="Filter by category">
    <button class="fchip on" data-cat="ALL">All 102</button>
    <button class="fchip" data-cat="UWC">UPVC Windows</button>
    <button class="fchip" data-cat="UWD">UPVC Doors</button>
    <button class="fchip" data-cat="ALW">Alu Windows</button>
    <button class="fchip" data-cat="ALD">Alu Doors</button>
    <button class="fchip" data-cat="ALP">Partitions</button>
    <button class="fchip" data-cat="GLZ">Glass</button>
    <button class="fchip" data-cat="STF">Storefront</button>
    <button class="fchip" data-cat="MST">MS Steel</button>
    <button class="fchip" data-cat="SLS">SS Steel</button>
    <button class="fchip" data-cat="RLS">Railings</button>
    <button class="fchip" data-cat="MOD">Modular</button>
  </div>
</div>
<div class="result-count" id="rcount">Showing 102 of 102 designs</div>

<main id="main" class="wrap">@@SECTIONS@@</main>

<section class="aboutsec" aria-label="About our workshop">
  <div class="wrap aboutgrid">
    <div class="reveal">
      <h2>Annapurna Aluminium &amp; UPVC — <em>premium quality, custom design</em></h2>
      <p>We are specialists in <strong>UPVC windows &amp; doors, aluminium windows, partitions, kitchen racks and steel railings</strong> — located at <strong>Janakpur Dham-8, Murli Chowk (Airport Road), Dhanusha (Madhesh Province)</strong>, with delivery &amp; installation across Nepal. Every product is factory-finished with premium profiles, branded hardware and precision fabrication.</p>
      <p>Visit our showroom to see 102+ ready designs or create your own custom design with free consultation. <strong>Competitive UPVC window prices in Nepal</strong> — call today for a free quote.</p>
      <ul class="aboutpoints">
        <li><span class="ic">&#10003;</span>Free site visit &amp; measurement — Janakpur, Dhanusha &amp; nearby districts</li>
        <li><span class="ic">&#10003;</span>German-profile UPVC — sound &amp; heat proof</li>
        <li><span class="ic">&#10003;</span>Toughened safety glass (as per safety norms)</li>
        <li><span class="ic">&#10003;</span>On-time delivery &amp; professional installation — Nepal-wide</li>
      </ul>
    </div>
    <div class="abouter reveal">
      <img src="@@WORKSHOP_IMG@@" alt="@@OWNER_NAME@@ in his UPVC window and aluminium door workshop — Janakpur Dham, Nepal" loading="lazy" decoding="async">
      <div class="ownercaption">
        <img src="@@AVATAR_CIRCLE@@" alt="" width="56" height="56">
        <div><b>@@OWNER_NAME@@</b><span>Founder &amp; Workshop Owner — Janakpur Dham</span></div>
      </div>
      <img class="boardimg" src="@@BOARD_IMG@@" alt="Annapurna Aluminium & UPVC shop board — Janakpur Dham-8, Murli Chowk, Nepal" loading="lazy" decoding="async">
    </div>
  </div>
</section>

<section class="areassec wrap" aria-label="Service areas in Nepal">
  <div class="cathead reveal">
    <span class="catnum">&#9733;</span>
    <span class="catico">&#9873;</span>
    <div>
      <h2>Service Areas — <em>Nepal-wide</em></h2>
      <p>From Janakpur Dham to every corner of Nepal — supply &amp; installation</p>
    </div>
  </div>
  <div class="areachips reveal">
    <span class="areachip main">Janakpur Dham</span>
    <span class="areachip">Dhanusha</span>
    <span class="areachip">Mahottari — Jaleshwor</span>
    <span class="areachip">Sarlahi — Malangwa</span>
    <span class="areachip">Siraha — Lahan</span>
    <span class="areachip">Sindhuli — Bardibas</span>
    <span class="areachip">Udayapur — Gaighat</span>
    <span class="areachip">Rajbiraj</span>
    <span class="areachip">Birgunj</span>
    <span class="areachip">Hetauda</span>
    <span class="areachip">Kathmandu Valley</span>
    <span class="areachip">Nepal-wide delivery</span>
  </div>
  <p class="areasnote reveal">Janakpur Dham, Dhanusha, Mahottari, Sarlahi, Siraha, Sindhuli, Udayapur, Bara, Parsa, Rautahat — <strong>free measurement &amp; professional installation available across all districts of Madhesh Province</strong>, with delivery to the rest of Nepal as well.</p>
</section>

<section class="faqsec wrap" aria-label="Frequently asked questions">
  <h2>Frequently Asked <em style="color:var(--gold3);font-style:italic">Questions</em></h2>
  <p class="lead">Straight answers to the questions our customers ask most</p>
  <div class="faqlist">@@FAQ_HTML@@</div>
</section>

<div class="banner" id="banner">
  <div class="bbg show" id="bbg0" role="img" aria-label="Premium design showcase"></div>
  <div class="bbg" id="bbg1" role="img" aria-hidden="true"></div>
  <div class="wrap">
    <h2>Every design is factory-finished<br>to your exact measurements.</h2>
    <p>Call us to book your order — free site visit, measurement, design consultation and installation, all with our team.</p>
    <a class="btn-gold" href="@@PHONE_LINK@@">&#9742; Call Now — @@PHONE@@</a>
  </div>
</div>

<section class="contact" id="contact" aria-label="Contact">
  <div class="wrap grid2">
    <div class="cbox"><h4>Visit / Call</h4>
      <p>@@ADDRESS_SHORT@@<br>Janakpur Dham-8, Dhanusha, Nepal<br>Call / WhatsApp: <a href="@@PHONE_LINK@@">@@PHONE@@</a><br>Alt: <a href="@@PHONE_LINK2@@">@@PHONE2@@</a><br>PAN No.: @@PAN@@</p></div>
    <div class="cbox"><h4>Services</h4>
      <p>Free measurement visit<br>Custom design consultation<br>Supply + installation<br>After-sales service &amp; warranty</p></div>
    <div class="cbox"><h4>Specialities</h4>
      <p>UPVC windows &amp; doors<br>Aluminium glazing &amp; partitions<br>Toughened / laminated glass<br>MS &amp; stainless steel works</p></div>
    <div class="cbox"><h4>Our Promise</h4>
      <p>Premium materials only<br>Factory-finished quality<br>On-time delivery<br>Fair pricing — no hidden cost</p></div>
    <div class="cbox ownerbox">
      <h4>Meet the Owner</h4>
      <div class="ownerrow">
        <img src="@@AVATAR_DARK@@" alt="@@OWNER_NAME@@ avatar" width="64" height="64">
        <div>
          <b>@@OWNER_NAME@@</b>
          <span>Founder &amp; Workshop Owner</span>
          <a href="@@PHONE_LINK@@">Call / WhatsApp — @@PHONE@@</a>
        </div>
      </div>
      <p class="ownertag">"Customer ki satisfaction hi hamari sabse badi advertisement hai."</p>
    </div>
  </div>
  <div class="wrap foot">
    <b>@@SHOP_NAME@@</b> — Premium UPVC windows (jhyal), aluminium doors (dhoka), partitions, kitchen racks &amp; steel railings in Janakpur Dham-8, Murli Chowk — Dhanusha, Madhesh Province, Nepal. Nepal-wide supply &amp; installation.<br>
    © 2026 @@SHOP_NAME@@ &nbsp;|&nbsp; Owner: @@OWNER_NAME@@ &nbsp;|&nbsp; PAN: @@PAN@@ &nbsp;|&nbsp; 102 unique premium designs &nbsp;|&nbsp; UPVC windows Nepal • Aluminium doors Janakpur • Glass works Dhanusha &nbsp;|&nbsp; Free measurement
  </div>
</section>

<a class="whatsapp" href="https://wa.me/@@WHATSAPP@@" target="_blank" rel="noopener" aria-label="WhatsApp enquiry">
  <svg viewBox="0 0 24 24"><path d="M12 2a10 10 0 0 0-8.6 15.1L2 22l5.1-1.3A10 10 0 1 0 12 2zm5.6 14.2c-.2.6-1.3 1.2-1.8 1.2-.5.1-1 .2-3.4-.7-2.9-1.2-4.7-4.1-4.9-4.3-.1-.2-1.1-1.5-1.1-2.9s.7-2 1-2.3c.2-.3.5-.3.7-.3h.5c.2 0 .4 0 .6.5l.9 2.1c.1.2.1.4 0 .6l-.4.6-.5.5c-.2.2-.3.4-.1.7.2.3.8 1.4 1.8 2.2 1.3 1.1 2.3 1.5 2.7 1.6.3.2.5.1.7-.1l1-1.2c.2-.3.4-.2.7-.1l2 1c.3.1.5.2.6.4 0 .1 0 .6-.3 1.2z"/></svg>
</a>
<button class="totop" aria-label="Back to top">&#8593;</button>

<div class="lightbox" id="lb" role="dialog" aria-label="Product photo preview">
  <button class="lbclose" id="lbclose" aria-label="Close">&#10005;</button>
  <img id="lbimg" src="" alt="Product enlarged view">
  <div class="lbmeta" id="lbmeta"></div>
</div>

@@JS@@
</body>
</html>"""

    doc = (doc.replace("@@FONTS@@", fonts_css)
              .replace("@@CSS@@", CSS)
              .replace("@@JS@@", JS.replace("@@HERO_URLS@@", json.dumps(hero_urls))
                                       .replace("@@BANNER_URLS@@", json.dumps(banner_urls)))
              .replace("@@NAV@@", nav)
              .replace("@@SECTIONS@@", "".join(sections))
              .replace("@@FAQ_HTML@@", faq_html)
              .replace("@@BIZ_JSON@@", build_biz_jsonld())
              .replace("@@ITEMLIST_JSON@@", build_itemlist_jsonld())
              .replace("@@FAQ_JSON@@", build_faq_jsonld())
              .replace("@@WEBSITE_JSON@@", build_website_jsonld())
              .replace("@@ABOUT_IMG@@", data_uri(ASSETS / "hero_facade.jpg", 760, 62))
              .replace("@@SHOP_NAME@@", SHOP["name"])
              .replace("@@BRAND@@", SHOP["brand"])
              .replace("@@DOMAIN@@", SHOP["domain"])
              .replace("@@PHONE@@", SHOP["phone"])
              .replace("@@PHONE2@@", SHOP["phone2"])
              .replace("@@PHONE_LINK@@", SHOP["phone_link"])
              .replace("@@PHONE_LINK2@@", "tel:+9779817667115")
              .replace("@@WHATSAPP@@", SHOP["whatsapp"])
              .replace("@@EMAIL@@", SHOP["email"])
              .replace("@@ADDRESS@@", SHOP["address"])
              .replace("@@ADDRESS_SHORT@@", SHOP["address_short"])
              .replace("@@PAN@@", SHOP["pan"])
              .replace("@@OWNER_NAME@@", SHOP["owner_name"])
              .replace("@@NAV_AVATAR@@", owner_uri("avatar1_circle.png", 90, 78) or "")
              .replace("@@AVATAR_GOLD@@", owner_uri("avatar2_goldring.png", 150) or "")
              .replace("@@AVATAR_CIRCLE@@", owner_uri("avatar1_circle.png", 120) or "")
              .replace("@@AVATAR_DARK@@", owner_uri("avatar4_darklux.png", 140) or "")
              .replace("@@MEET_IMG@@", owner_uri("owner_meet.jpg", 620, 74) or "")
              .replace("@@CHAR_IMG@@", owner_uri("avatar7_character.png", 240, 80) or "")
              .replace("@@WORKSHOP_IMG@@", owner_uri("owner_workshop.jpg", 640, 72) or "")
              .replace("@@BOARD_IMG@@", owner_uri("shop_board.jpg", 720, 74) or "")
              .replace("@@FAVICON@@", favicon))
    (OUT / "catalog.html").write_text(doc, encoding="utf-8")
    print("catalog.html written:", round(len(doc) / 1024 / 1024, 2), "MB")

    # ---------------- POSTER (matching premium style) ----------------
    picks = []
    for c in CATS:
        picks += by_cat[c["key"]][:3]
    pcards = "".join(f"""
      <div class="pcard">
        <div class="part">{('<img src="%s" alt="%s" loading="lazy">' % (photo_uri(pr['code'], 520, 62), esc(pr['name']))) if photo_uri(pr['code']) else svg_wrap(art_map[pr['code']])}</div>
        <div class="pbody"><span class="pcode">{esc(pr['code'])}</span>
        <b>{esc(pr['name'])}</b><p>{esc(pr['desc'])}</p></div>
      </div>""" for pr in picks)
    poster = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Showcase Poster — Alu • Steel • UPVC</title>
<style>@@FONTS@@
body{font-family:'Manrope','Segoe UI',Arial,sans-serif;background:#0d1526;color:#eef3f8;margin:0}
.pwrap{width:min(1200px,96%);margin:0 auto;padding:36px 0 46px}
.ptop{text-align:center;margin-bottom:30px}
.ptop h1{font-family:'Playfair Display',Georgia,serif;font-size:46px;margin:0 0 6px;letter-spacing:.02em;color:#fff}
.ptop h1 span{color:#f3c25b;font-style:italic}
.ptop p{color:#9fb0c2;letter-spacing:.26em;font-size:11px;text-transform:uppercase}
.pgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(255px,1fr));gap:14px}
.pcard{background:#182130;border:1px solid #2a384a;border-radius:12px;overflow:hidden}
.part{background:linear-gradient(160deg,#1c2635,#121a26);padding:0}
.part img{display:block;width:100%;height:150px;object-fit:cover}
.part svg{display:block;width:100%;height:auto;padding:10px 12px 6px}
.pbody{padding:11px 14px 14px}
.pcode{font-size:10px;font-weight:800;letter-spacing:.12em;color:#f3c25b}
.pbody b{font-family:'Playfair Display',Georgia,serif;display:block;font-size:14.5px;margin:3px 0 4px}
.pbody p{color:#9fb0c2;font-size:11.5px;line-height:1.45;margin:0}
.pfoot{text-align:center;color:#6d7f94;font-size:12px;margin-top:26px;letter-spacing:.14em}
.poster-owner{display:flex;align-items:center;gap:16px;justify-content:center;flex-wrap:wrap;margin-top:28px;
  background:#182130;border:1px solid #2a384a;border-radius:16px;padding:16px 24px}
.poster-owner img{width:62px;height:62px;border-radius:50%;border:2.5px solid #f3c25b;object-fit:cover}
.po-info b{display:block;color:#fff;font-family:'Playfair Display',Georgia,serif;font-size:17px}
.po-info span{color:#f3c25b;font-size:10.5px;letter-spacing:.12em;text-transform:uppercase;font-weight:700}
.po-call{text-align:right;border-left:1px solid #2a384a;padding-left:18px}
.po-call span{color:#9fb0c2;font-size:11px;display:block}
.po-call b{color:#f3c25b;font-size:16px}
@media print{body{background:#fff}.pcard{break-inside:avoid;background:#fff;border-color:#ccc}.pbody p,.pfoot{color:#555}.ptop p{color:#666}}
</style></head><body>
<div class="pwrap">
  <div class="ptop"><h1>Alu <span>•</span> Steel <span>•</span> UPVC</h1>
  <p>Premium Glass &amp; Metal Works — Showcase Collection</p></div>
  <div class="pgrid">@@PCARDS@@</div>
  <div class="poster-owner">
    <img src="@@POSTER_AVATAR@@" alt="@@OWNER_NAME@@ — owner" width="62" height="62">
    <div class="po-info">
      <b>@@OWNER_NAME@@</b>
      <span>Founder &amp; Workshop Owner — Janakpur Dham, Nepal</span>
    </div>
    <div class="po-call">
      <span>Free Measurement &amp; Quote</span>
      <b>&#9742; @@PHONE@@</b>
    </div>
  </div>
  <p class="pfoot">102 UNIQUE DESIGNS • CUSTOM SIZES • FREE MEASUREMENT • CALL: @@PHONE@@</p>
</div></body></html>"""
    poster = (poster.replace("@@FONTS@@", fonts_css)
                    .replace("@@PCARDS@@", pcards)
                    .replace("@@PHONE@@", SHOP["phone"])
                    .replace("@@OWNER_NAME@@", SHOP["owner_name"])
                    .replace("@@POSTER_AVATAR@@", owner_uri("avatar2_goldring.png", 140, 78) or ""))
    (OUT / "poster.html").write_text(poster, encoding="utf-8")
    print("poster.html written:", round(len(poster) / 1024, 1), "KB")

# ---------------- XLSX ----------------
def build_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"
    thin = Side(style="thin", color="D8DEE6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def put(r, c, v, font=None, fill=None, align=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or Font(size=10)
        if fill: cell.fill = fill
        if align: cell.alignment = align
        cell.border = border
        return cell

    ws.merge_cells("A1:K1")
    put(1, 1, "PREMIUM PRODUCT SHOWCASE — PRICE LIST  |  %s" % SHOP["short"], Font(size=16, bold=True, color="FFFFFF"),
        PatternFill("solid", fgColor="0B1220"), Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:K2")
    put(2, 1, "Aluminium • Steel • UPVC Glass Works — Janakpur Dham, Dhanusha, Madhesh Province, Nepal  |  Custom Sizes  |  Premium Quality  |  Free Measurement  |  %s" % SHOP["phone"],
        Font(size=10, italic=True, color="7A5B12"), PatternFill("solid", fgColor="F7E8C3"), Alignment(horizontal="center"))
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A3:K3")
    put(3, 1, "Prices: On Request — fill in as per your workshop rate card.  |  All designs factory-finished & installation-ready.",
        Font(size=9, color="5B6B7D"), None, Alignment(horizontal="center"))
    ws.row_dimensions[3].height = 18
    ws.merge_cells("A4:K4")
    put(4, 1, "Owner: %s  |  Call / WhatsApp: %s  |  Address: %s" % (SHOP["owner_name"], SHOP["phone"], SHOP["address"]),
        Font(size=9.5, bold=True, color="FFFFFF"), PatternFill("solid", fgColor="3A4A5C"),
        Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[4].height = 18

    headers = ["S.No", "Code", "Category", "Product Name", "Description", "Standard Sizes",
               "Finishes", "Key Features", "Price (NRs)", "Units", "Status"]
    hr = 5
    for ci, h in enumerate(headers, 1):
        put(hr, ci, h, Font(size=10, bold=True, color="FFFFFF"), PatternFill("solid", fgColor="243447"),
            Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[hr].height = 26

    for i, w in enumerate([6, 10, 15, 26, 42, 20, 20, 44, 12, 8, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = hr + 1
    sno = 0
    by_cat = {}
    for pr in P:
        by_cat.setdefault(pr["cat"], []).append(pr)
    for c in CATS:
        items = by_cat[c["key"]]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        put(r, 1, "%s — %s  (%d designs)" % (c["key"], c["name"], len(items)),
            Font(size=10.5, bold=True, color="FFFFFF"), PatternFill("solid", fgColor="B98A2F"),
            Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[r].height = 20
        r += 1
        for pr in items:
            sno += 1
            vals = [sno, pr["code"], pr["cat"], pr["name"], pr["desc"], "; ".join(pr["sizes"]),
                    "; ".join(pr["finishes"]), "; ".join(pr["extras"]), "On Request", "Set", "Available"]
            for ci, v in enumerate(vals, 1):
                al = Alignment(vertical="top", wrap_text=(ci in (4, 5, 6, 7, 8)),
                               horizontal="center" if ci in (1, 2, 9, 10, 11) else "left")
                put(r, ci, v, Font(size=9.5, bold=(ci == 4)),
                    fill=(PatternFill("solid", fgColor="F4F6F9") if sno % 2 == 0 else None), align=al)
            ws.row_dimensions[r].height = 52
            r += 1
    ws.freeze_panes = "A6"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(OUT / "price_list.xlsx")
    print("price_list.xlsx written —", len(P), "products")

# ---------------- run ----------------
if __name__ == "__main__":
    from urllib.parse import quote as urllib_quote
    art = render_art()
    missing = [p["code"] for p in P if not art.get(p["code"])]
    if missing:
        print("MISSING ART:", missing)
    build_html(art)
    build_xlsx()
    print("DONE — total designs:", len(P))
